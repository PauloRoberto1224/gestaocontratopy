"""
Utility functions and helpers for the project.
"""
import csv
import json
import logging
import re
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Tuple, Union

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import QuerySet, Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.encoding import force_str
from django.utils.html import escape
from django.utils.text import slugify
from django.utils.translation import gettext_lazy as _

logger = logging.getLogger(__name__)

# Date/Time Utilities

def now() -> datetime:
    """Return current time in the default timezone."""
    return timezone.now()

today = now().date()

def format_date(value: Union[date, datetime, str], fmt: str = 'd/m/Y') -> str:
    """Format a date or datetime object as string."""
    if not value:
        return ''
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return value
    return value.strftime(fmt)

def parse_date(date_str: str, fmt: str = '%d/%m/%Y') -> Optional[date]:
    """Parse a date string into a date object."""
    try:
        return datetime.strptime(date_str, fmt).date()
    except (ValueError, TypeError):
        return None

# String Utilities

def slugify_text(text: str) -> str:
    """Convert text to a URL-friendly slug."""
    if not text:
        return ''
    text = force_str(text)
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    return re.sub(r'[-\s]+', '-', text)

# Number/Currency Utilities

def format_currency(value: Union[Decimal, float, int, str], currency: str = 'R$') -> str:
    """Format a number as currency."""
    if value is None:
        return f"{currency} 0,00"
    try:
        if isinstance(value, str):
            value = value.replace('.', '').replace(',', '.')
            value = Decimal(value)
        elif isinstance(value, (int, float)):
            value = Decimal(str(value))
        
        return f"{currency} {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError, InvalidOperation) as e:
        logger.warning(f"Error formatting currency: {e}")
        return f"{currency} 0,00"

# QuerySet Utilities

def get_object_or_none(klass, *args, **kwargs):
    """Get an object or return None if it doesn't exist."""
    try:
        return klass._default_manager.get(*args, **kwargs)
    except (klass.DoesNotExist, ValidationError):
        return None

def paginate_queryset(queryset: QuerySet, request: HttpRequest, page_size: int = 20) -> dict:
    """Paginate a queryset and return pagination data."""
    paginator = Paginator(queryset, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    return {
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': page_obj.has_other_pages(),
    }

# Export Utilities

def export_to_csv(queryset, fields, filename='export.csv'):
    """Export a queryset to CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([field[0] for field in fields])  # Headers
    
    for obj in queryset:
        row = []
        for _, attr in fields:
            value = getattr(obj, attr, '')
            if callable(value):
                value = value()
            row.append(str(value) if value is not None else '')
        writer.writerow(row)
    
    return response

def export_to_excel(queryset, fields, filename='export.xlsx'):
    """Export a queryset to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Export'
    
    # Headers
    for col_num, (header, _) in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, attr) in enumerate(fields, 1):
            value = getattr(obj, attr, '')
            if callable(value):
                value = value()
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    wb.save(response)
    return response

# Request/Response Utilities

def get_client_ip(request: HttpRequest) -> str:
    """Get the client's IP address."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR', '')

def get_next_or_prev(queryset, current, direction='next'):
    """
    Get the next or previous object in a queryset.
    
    Args:
        queryset: The queryset to search in
        current: The current object
        direction: 'next' or 'prev'
    
    Returns:
        The next or previous object, or None if not found
    """
    if not current or not hasattr(current, 'id'):
        return None
        
    filter_kwargs = {}
    order_by = 'id'
    
    if hasattr(current, 'created_at'):
        order_by = 'created_at'
    
    if direction == 'next':
        filter_kwargs[f'{order_by}__gt'] = getattr(current, order_by)
        return queryset.filter(**filter_kwargs).order_by(order_by).first()
    else:  # prev
        filter_kwargs[f'{order_by}__lt'] = getattr(current, order_by)
        return queryset.filter(**filter_kwargs).order_by(f'-{order_by}').first()

def get_user_agent(request):
    """
    Get user agent information from the request.
    
    Args:
        request: The HTTP request object
        
    Returns:
        A dictionary with user agent information
    """
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    # Basic user agent parsing
    is_mobile = any(m in user_agent.lower() for m in [
        'mobile', 'android', 'iphone', 'ipad', 'ipod', 'blackberry', 
        'windows phone', 'opera mini', 'iemobile'
    ])
    
    is_tablet = any(t in user_agent.lower() for t in [
        'ipad', 'tablet', 'playbook', 'kindle', 'silk', 'xoom'
    ])
    
    browser = 'Unknown'
    if 'firefox' in user_agent.lower():
        browser = 'Firefox'
    elif 'chrome' in user_agent.lower() and 'edg' not in user_agent.lower():
        browser = 'Chrome'
    elif 'safari' in user_agent.lower() and 'chrome' not in user_agent.lower():
        browser = 'Safari'
    elif 'edg' in user_agent.lower():
        browser = 'Edge'
    elif 'msie' in user_agent.lower() or 'trident' in user_agent.lower():
        browser = 'Internet Explorer'
    
    os = 'Unknown'
    if 'windows' in user_agent.lower():
        os = 'Windows'
    elif 'mac os x' in user_agent.lower():
        os = 'macOS'
    elif 'linux' in user_agent.lower() and 'android' not in user_agent.lower():
        os = 'Linux'
    elif 'android' in user_agent.lower():
        os = 'Android'
    elif 'iphone' in user_agent.lower() or 'ipad' in user_agent.lower() or 'ipod' in user_agent.lower():
        os = 'iOS'
    
    return {
        'user_agent': user_agent,
        'is_mobile': is_mobile,
        'is_tablet': is_tablet,
        'is_desktop': not (is_mobile or is_tablet),
        'browser': browser,
        'os': os
    }

def send_email(subject, message, to_emails, from_email=None, html_message=None, **kwargs):
    """
    Send an email using Django's send_mail function.
    
    Args:
        subject (str): Email subject
        message (str): Plain text email content
        to_emails (list): List of recipient email addresses
        from_email (str, optional): Sender email address. Defaults to DEFAULT_FROM_EMAIL.
        html_message (str, optional): HTML version of the email content.
        **kwargs: Additional arguments to pass to send_mail
    
    Returns:
        int: Number of emails sent
    """
    from django.core.mail import send_mail as django_send_mail
    from django.conf import settings
    
    if not from_email:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'webmaster@localhost')
    
    if not isinstance(to_emails, (list, tuple)):
        to_emails = [to_emails]
    
    return django_send_mail(
        subject=subject,
        message=message,
        from_email=from_email,
        recipient_list=to_emails,
        html_message=html_message,
        **kwargs
    )

def generate_pdf(content, filename='document.pdf', **kwargs):
    """
    Generate a PDF file from HTML content.
    
    Args:
        content (str): HTML content to convert to PDF
        filename (str): Output filename
        **kwargs: Additional arguments for PDF generation
        
    Returns:
        HttpResponse: PDF file response
    """
    from django.http import HttpResponse
    from xhtml2pdf import pisa
    from io import BytesIO
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    pdf = pisa.CreatePDF(BytesIO(content.encode('UTF-8')), response, **kwargs)
    
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)
        
    return response

def export_to_pdf(queryset, fields, filename='export.pdf', **kwargs):
    """
    Export a queryset to PDF.
    
    Args:
        queryset: The queryset to export
        fields: List of (header, field_name) tuples
        filename: Output filename
        **kwargs: Additional arguments for PDF generation
        
    Returns:
        HttpResponse: PDF file response
    """
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    
    context = {
        'objects': queryset,
        'fields': fields,
        'title': 'Relatório de Contratos',
        'date': timezone.now().strftime('%d/%m/%Y %H:%M'),
    }
    
    html_string = render_to_string('reports/pdf_template.html', context)
    return generate_pdf(html_string, filename=filename, **kwargs)

def format_datetime(value, fmt='%d/%m/%Y %H:%M'):
    """Format a datetime object as string."""
    if not value:
        return ''
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            return value
    return value.strftime(fmt)

def format_time(value, fmt='%H:%M'):
    """Format a time object as string."""
    if not value:
        return ''
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%H:%M:%S').time()
        except (ValueError, TypeError):
            return value
    return value.strftime(fmt)

def format_phone(phone):
    """Format a phone number."""
    if not phone:
        return ''
    
    phone = ''.join(filter(str.isdigit, str(phone)))
    
    if len(phone) == 11:  # With DDD and 9th digit
        return f'({phone[:2]}) {phone[2:7]}-{phone[7:]}'
    elif len(phone) == 10:  # With DDD
        return f'({phone[:2]}) {phone[2:6]}-{phone[6:]}'
    elif len(phone) == 9:  # 9th digit only
        return f'{phone[:5]}-{phone[5:]}'
    elif len(phone) == 8:  # Traditional number
        return f'{phone[:4]}-{phone[4:]}'
    return phone

def get_paginator_range(page_obj, on_each_side=3, on_ends=2):
    """
    Generate a list of page numbers for pagination.
    
    Args:
        page_obj: The current page object from Django's paginator
        on_each_side: Number of pages to show on each side of the current page
        on_ends: Number of pages to show at the start and end
        
    Returns:
        list: List of page numbers with None for ellipsis
    """
    current_page = page_obj.number
    total_pages = page_obj.paginator.num_pages
    
    if total_pages <= (on_each_side + on_ends) * 2 + 1:
        return list(range(1, total_pages + 1))
    
    pages = []
    
    # Add pages at the start
    pages.extend(range(1, on_ends + 1))
    
    # Add ellipsis if needed
    if current_page - on_each_side - 1 > on_ends:
        pages.append(None)
    
    # Add pages around current page
    start = max(on_ends + 1, current_page - on_each_side)
    end = min(total_pages - on_ends, current_page + on_each_side)
    pages.extend(range(start, end + 1))
    
    # Add ellipsis if needed
    if current_page + on_each_side + 1 < total_pages - on_ends + 1:
        pages.append(None)
    
    # Add pages at the end
    if total_pages - on_ends + 1 > 0:
        pages.extend(range(total_pages - on_ends + 1, total_pages + 1))
    
    return pages

def get_ordering(request, default_ordering='-created_at'):
    """
    Get the ordering parameter from the request.
    
    Args:
        request: The HTTP request object
        default_ordering: Default ordering if none is specified
        
    Returns:
        str: Ordering parameter
    """
    return request.GET.get('order_by', default_ordering)

def get_search_fields():
    """
    Get the default search fields for models.
    
    Returns:
        list: List of default search fields
    """
    return ['name__icontains', 'description__icontains']

def get_filtered_queryset(queryset, request, search_fields=None, filters=None):
    """
    Apply search and filter to a queryset.
    
    Args:
        queryset: The base queryset
        request: The HTTP request object
        search_fields: Fields to search in
        filters: Additional filters to apply
        
    Returns:
        QuerySet: Filtered queryset
    """
    from django.db.models import Q
    
    # Apply search
    search_query = request.GET.get('q', '').strip()
    if search_query and search_fields:
        search_query = search_query.split()
        search_filter = Q()
        
        for term in search_query:
            term_filter = Q()
            for field in search_fields:
                term_filter |= Q(**{field: term})
            search_filter &= term_filter
            
        queryset = queryset.filter(search_filter)
    
    # Apply additional filters
    if filters:
        filter_kwargs = {}
        for param, field in filters.items():
            value = request.GET.get(param)
            if value:
                filter_kwargs[field] = value
        
        if filter_kwargs:
            queryset = queryset.filter(**filter_kwargs)
    
    return queryset

def get_ordered_queryset(queryset, request, default_ordering='-created_at'):
    """
    Apply ordering to a queryset.
    
    Args:
        queryset: The base queryset
        request: The HTTP request object
        default_ordering: Default ordering if none is specified
        
    Returns:
        QuerySet: Ordered queryset
    """
    ordering = get_ordering(request, default_ordering)
    if ordering:
        return queryset.order_by(ordering)
    return queryset

def get_paginated_data(queryset, request, per_page=20):
    """
    Paginate a queryset.
    
    Args:
        queryset: The base queryset
        request: The HTTP request object
        per_page: Number of items per page
        
    Returns:
        dict: Pagination data including page object and page range
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(queryset, per_page)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    page_range = get_paginator_range(page_obj)
    
    return {
        'page_obj': page_obj,
        'paginator': paginator,
        'page_range': page_range,
        'is_paginated': page_obj.has_other_pages(),
    }