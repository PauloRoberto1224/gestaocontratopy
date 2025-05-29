import csv
import json
from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Union, Any

# Django imports
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User, Group
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import (
    Q, Count, Sum, Avg, F, Value, Case, When, IntegerField,
    CharField, TextField, DateField, DateTimeField, DecimalField,
    BooleanField, ForeignKey, ManyToManyField, JSONField
)
from django.db.models.functions import (
    Concat, Coalesce, TruncDate, TruncMonth, TruncYear,
    ExtractYear, ExtractMonth, ExtractDay, ExtractWeekDay,
    Now, Lower, Upper, Substr, Length
)
from django.http import (
    HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse,
    HttpResponseForbidden, HttpResponseBadRequest, HttpResponseServerError,
    FileResponse, Http404, StreamingHttpResponse
)
from django.db.models import Q
from django.shortcuts import (
    render, redirect, get_object_or_404, get_list_or_404
)
from django.template.loader import render_to_string, get_template
from django.urls import reverse_lazy, reverse, resolve, NoReverseMatch
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.encoding import force_str, force_bytes
from django.utils.formats import date_format, number_format
from django.utils.html import strip_tags, escape, format_html, format_html_join
from django.utils.safestring import mark_safe
from django.utils.text import slugify, capfirst, get_text_list
from django.utils.timezone import localtime, now, make_aware, is_naive
from django.utils.translation import gettext_lazy as _, get_language, activate
from django.views import View
from django.views.decorators.cache import cache_page, never_cache
from django.views.decorators.csrf import csrf_exempt, csrf_protect, ensure_csrf_cookie
from django.views.decorators.http import (
    require_http_methods, require_GET, require_POST, require_safe, require_http_methods
)
from django.views.generic import (
    View, TemplateView, RedirectView,
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    ArchiveIndexView, YearArchiveView, MonthArchiveView, DayArchiveView,
    TodayArchiveView, DateDetailView, WeekArchiveView
)
from django.views.generic.edit import FormView, FormMixin
from django.views.generic.base import ContextMixin, TemplateResponseMixin
from django.views.generic.detail import SingleObjectMixin
from django.views.generic.list import MultipleObjectMixin

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle, StyleSheet1
from reportlab.lib.units import inch, cm, mm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    ListFlowable, ListItem, PageBreak, NextPageTemplate, Frame, PageTemplate,
    KeepTogether, PageBreakIfNotEmpty, HRFlowable, KeepInFrame, FrameBreak,
    ListItem, ListFlowable, PageBreak, NextPageTemplate, Frame, PageTemplate,
    KeepTogether, PageBreakIfNotEmpty, HRFlowable, KeepInFrame, FrameBreak
)
from reportlab.platypus.flowables import KeepTogether, Spacer, HRFlowable, PageBreak
from reportlab.platypus.frames import Frame
from reportlab.platypus.para import Paragraph
from reportlab.platypus.tables import Table, TableStyle
from reportlab.platypus.xpreformatted import XPreformatted
from reportlab.rl_config import defaultPageSize

# Local application imports
from core.mixins import StaffRequiredMixin, SuperuserRequiredMixin, AjaxResponseMixin, JSONResponseMixin
from core.utils import (
    get_object_or_none, get_next_or_prev, get_client_ip, get_user_agent,
    send_email, generate_pdf, export_to_excel, export_to_csv, export_to_pdf,
    format_currency, format_date, format_datetime, format_time, format_phone,
    paginate_queryset, get_paginator_range, get_ordering, get_search_fields,
    get_filtered_queryset, get_ordered_queryset, get_paginated_data
)

# Import models and forms
from .models import (
    Contract, ContractType, ContractStatus, ContractParty,
    ContractHistory, ContractReminder, ContractAttachment
)

from .forms import (
    ContractForm, ContractTypeForm, ContractStatusForm,
    ContractPartyForm, ContractReminderForm, ContractAttachmentForm,
    ContractFilterForm
)


# Relatórios
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import (
    Contract, ContractType, ContractStatus, ContractParty,
    ContractHistory, ContractReminder, ContractAttachment
)
from .forms import (
    ContractForm, ContractTypeForm, ContractStatusForm,
    ContractPartyForm, ContractReminderForm, ContractAttachmentForm
)
from core.mixins import StaffRequiredMixin


class ContractListView(LoginRequiredMixin, ListView):
    model = Contract
    template_name = 'contracts/contract_list.html'
    context_object_name = 'contracts'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Contract.objects.select_related(
            'status', 'responsible', 'created_by'
        )
        
        # Filtros
        self.filter_form = ContractFilterForm(self.request.GET or None)
        if self.filter_form.is_valid():
            data = self.filter_form.cleaned_data
            
            # Filtro por texto
            if data.get('search'):
                search = data['search']
                queryset = queryset.filter(
                    Q(contract_number__icontains=search) |
                    Q(company__icontains=search) |
                    Q(fiscal_name__icontains=search) |
                    Q(notes__icontains=search)
                )
            
            # Filtro por termo do contrato
            if data.get('contract_term'):
                queryset = queryset.filter(contract_term=data['contract_term'])
            
            # Filtro por status
            if data.get('status'):
                queryset = queryset.filter(status=data['status'])
            
            # Filtro por responsável
            if data.get('responsible'):
                queryset = queryset.filter(responsible=data['responsible'])
            
            # Filtro por data de vencimento
            if data.get('expires_in') == 'week':
                next_week = timezone.now().date() + timezone.timedelta(days=7)
                queryset = queryset.filter(
                    end_date__lte=next_week,
                    end_date__gte=timezone.now().date()
                )
            elif data.get('expires_in') == 'month':
                next_month = timezone.now().date() + timezone.timedelta(days=30)
                queryset = queryset.filter(
                    end_date__lte=next_month,
                    end_date__gte=timezone.now().date()
                )
            elif data.get('expires_in') == 'expired':
                queryset = queryset.filter(end_date__lt=timezone.now().date())
        
        # Ordenação
        order_by = self.request.GET.get('order_by', '-created_at')
        if order_by.lstrip('-') in [f.name for f in Contract._meta.fields]:
            queryset = queryset.order_by(order_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = self.filter_form
        # Add contract term choices to the context for the template
        context['contract_term_choices'] = Contract.TERM_CHOICES
        context['statuses'] = ContractStatus.objects.all()
        return context


class ContractDetailView(LoginRequiredMixin, DetailView):
    model = Contract
    template_name = 'contracts/contract_detail.html'
    context_object_name = 'contract'
    
    def get_queryset(self):
        return Contract.objects.select_related(
            'status', 'responsible', 'created_by'
        ).prefetch_related('contract_parties__user')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['reminders'] = self.object.reminders.all().order_by('due_date')
        context['history'] = self.object.history_entries.all().order_by('-change_date')[:10]
        return context


class ContractCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Contract
    form_class = ContractForm
    template_name = 'contracts/contract_form.html'
    permission_required = 'contracts.add_contract'
    
    def get_success_url(self):
        return reverse_lazy('contracts:contract_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        # Processa o valor monetário antes de salvar
        if 'value' in form.cleaned_data and form.cleaned_data['value']:
            try:
                # Remove formatação do valor (R$, pontos e troca vírgula por ponto)
                value = form.cleaned_data['value']
                if isinstance(value, str):
                    value = value.replace('R$', '').replace('.', '').replace(',', '.').strip()
                    form.cleaned_data['value'] = value
            except (ValueError, AttributeError):
                pass
        
        form.instance.created_by = self.request.user
        
        try:
            response = super().form_valid(form)
            
            # Registrar histórico
            ContractHistory.objects.create(
                contract=self.object,
                changed_by=self.request.user,
                change_description=_('Contrato criado'),
                changed_fields={'created': True}
            )
            
            is_ajax = self.request.POST.get('is_ajax') == '1' or self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if is_ajax:
                data = {
                    'success': True,
                    'redirect': self.get_success_url(),
                    'message': _('Contrato criado com sucesso!'),
                    'contract_id': self.object.id
                }
                return JsonResponse(data)
            
            messages.success(self.request, _('Contrato criado com sucesso!'))
            return response
            
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': str(e),
                    'errors': {'__all__': [str(e)]}
                }, status=400)
            raise
    
    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors.get_json_data(),
                'message': _('Por favor, corrija os erros abaixo.')
            }, status=400)
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Novo Contrato')
        return context


class ContractUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Contract
    form_class = ContractForm
    template_name = 'contracts/contract_form.html'
    permission_required = 'contracts.change_contract'
    
    def get_success_url(self):
        return reverse_lazy('contracts:contract_detail', kwargs={'pk': self.object.pk})
    
    def get_queryset(self):
        return Contract.objects.all()
    
    def form_valid(self, form):
        # Salvar os dados antigos para comparação
        old_instance = Contract.objects.get(pk=self.object.pk)
        old_data = {
            'title': old_instance.title,
            'description': old_instance.description,
            'start_date': str(old_instance.start_date) if old_instance.start_date else None,
            'end_date': str(old_instance.end_date) if old_instance.end_date else None,
            'value': str(old_instance.value) if old_instance.value is not None else None,
            'status': str(old_instance.status) if old_instance.status else None,
            'contract_type': str(old_instance.contract_type) if old_instance.contract_type else None,
        }
        
        response = super().form_valid(form)
        
        # Obter os novos dados
        new_data = {
            'title': self.object.title,
            'description': self.object.description,
            'start_date': str(self.object.start_date) if self.object.start_date else None,
            'end_date': str(self.object.end_date) if self.object.end_date else None,
            'value': str(self.object.value) if self.object.value is not None else None,
            'status': str(self.object.status) if self.object.status else None,
            'contract_type': str(self.object.contract_type) if self.object.contract_type else None,
        }
        
        # Verificar alterações
        changed_fields = {}
        for field, old_value in old_data.items():
            if old_value != new_data[field]:
                changed_fields[field] = {
                    'old': old_value,
                    'new': new_data[field]
                }
        
        # Registrar histórico se houver alterações
        if changed_fields:
            ContractHistory.objects.create(
                contract=self.object,
                changed_by=self.request.user,
                change_description=_('Dados do contrato atualizados'),
                changed_fields=changed_fields
            )
        
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = {
                'success': True,
                'redirect': self.get_success_url(),
                'message': _('Contrato atualizado com sucesso!')
            }
            return JsonResponse(data)
        
        messages.success(self.request, _('Contrato atualizado com sucesso!'))
        return response
    
    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors.get_json_data(),
                'message': _('Por favor, corrija os erros abaixo.')
            }, status=400)
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Editar Contrato')
        return context


class ContractDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Contract
    template_name = 'contracts/contract_confirm_delete.html'
    success_url = reverse_lazy('contracts:contract_list')
    permission_required = 'contracts.delete_contract'
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Contrato excluído com sucesso!'))
        return super().delete(request, *args, **kwargs)


# Visualizações para Tipos de Contrato
class ContractTypeListView(LoginRequiredMixin, ListView):
    model = ContractType
    template_name = 'contracts/contracttype_list.html'
    context_object_name = 'types'
    
    def get_queryset(self):
        return ContractType.objects.annotate(
            contract_count=Count('contracts')
        ).order_by('name')


class ContractTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ContractType
    form_class = ContractTypeForm
    template_name = 'contracts/contracttype_form.html'
    permission_required = 'contracts.add_contracttype'
    success_url = reverse_lazy('contracts:contract_type_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Tipo de contrato criado com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Novo Tipo de Contrato')
        return context


class ContractTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ContractType
    form_class = ContractTypeForm
    template_name = 'contracts/contracttype_form.html'
    permission_required = 'contracts.change_contracttype'
    success_url = reverse_lazy('contracts:contract_type_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Tipo de contrato atualizado com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Editar Tipo de Contrato')
        return context


class ContractTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ContractType
    template_name = 'contracts/contracttype_confirm_delete.html'
    permission_required = 'contracts.delete_contracttype'
    success_url = reverse_lazy('contracts:contract_type_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Tipo de contrato excluído com sucesso!'))
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Excluir Tipo de Contrato')
        return context


# Visualizações para Status de Contrato
class ContractStatusListView(LoginRequiredMixin, ListView):
    model = ContractStatus
    template_name = 'contracts/contractstatus_list.html'
    context_object_name = 'status_list'
    
    def get_queryset(self):
        return ContractStatus.objects.annotate(
            contract_count=Count('contracts')
        ).order_by('name')


class ContractStatusCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ContractStatus
    form_class = ContractStatusForm
    template_name = 'contracts/contractstatus_form.html'
    permission_required = 'contracts.add_contractstatus'
    success_url = reverse_lazy('contracts:contract_status_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Status de contrato criado com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Novo Status de Contrato')
        return context


class ContractStatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ContractStatus
    form_class = ContractStatusForm
    template_name = 'contracts/contractstatus_form.html'
    permission_required = 'contracts.change_contractstatus'
    success_url = reverse_lazy('contracts:contract_status_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Status de contrato atualizado com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Editar Status de Contrato')
        return context


class ContractStatusDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ContractStatus
    template_name = 'contracts/contractstatus_confirm_delete.html'
    permission_required = 'contracts.delete_contractstatus'
    success_url = reverse_lazy('contracts:contract_status_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Status de contrato excluído com sucesso!'))
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Excluir Status de Contrato')
        return context


# Visualizações para Partes Contratuais
class ContractPartyListView(LoginRequiredMixin, ListView):
    model = ContractParty
    template_name = 'contracts/contractparty_list.html'
    context_object_name = 'parties'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = ContractParty.objects.all()
        
        # Filtro por nome ou documento
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(document_number__icontains=search)
            )
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context


class ContractPartyCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ContractParty
    form_class = ContractPartyForm
    template_name = 'contracts/contractparty_form.html'
    permission_required = 'contracts.add_contractparty'
    success_url = reverse_lazy('contracts:contract_party_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Parte contratual criada com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Nova Parte Contratual')
        return context


class ContractPartyUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ContractParty
    form_class = ContractPartyForm
    template_name = 'contracts/contractparty_form.html'
    permission_required = 'contracts.change_contractparty'
    success_url = reverse_lazy('contracts:contract_party_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Parte contratual atualizada com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Editar Parte Contratual')
        return context


class ContractPartyDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ContractParty
    template_name = 'contracts/contractparty_confirm_delete.html'
    permission_required = 'contracts.delete_contractparty'
    success_url = reverse_lazy('contracts:contract_party_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Parte contratual excluída com sucesso!'))
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Excluir Parte Contratual')
        return context


# Visualizações para Lembretes de Contrato
class ContractReminderListView(LoginRequiredMixin, ListView):
    model = ContractReminder
    template_name = 'contracts/contractreminder_list.html'
    context_object_name = 'reminders'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = ContractReminder.objects.select_related('contract', 'created_by')
        
        # Filtro por status (pendente/realizado)
        status = self.request.GET.get('status')
        if status == 'pending':
            queryset = queryset.filter(completed=False)
        elif status == 'completed':
            queryset = queryset.filter(completed=True)
        
        # Filtro por data
        due_date = self.request.GET.get('due_date')
        if due_date:
            queryset = queryset.filter(due_date=due_date)
        
        # Filtro por contrato
        contract_id = self.request.GET.get('contract')
        if contract_id:
            queryset = queryset.filter(contract_id=contract_id)
        
        return queryset.order_by('due_date', 'contract__title')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Adicionar filtros atuais ao contexto
        context['status'] = self.request.GET.get('status', '')
        context['due_date'] = self.request.GET.get('due_date', '')
        context['contract_id'] = self.request.GET.get('contract', '')
        
        # Adicionar lista de contratos para o filtro
        context['contracts'] = Contract.objects.all().order_by('title')
        
        # Estatísticas
        today = timezone.now().date()
        context['pending_count'] = ContractReminder.objects.filter(completed=False).count()
        context['overdue_count'] = ContractReminder.objects.filter(
            completed=False,
            due_date__lt=today
        ).count()
        
        return context


class ContractReminderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ContractReminder
    form_class = ContractReminderForm
    template_name = 'contracts/contractreminder_form.html'
    permission_required = 'contracts.add_contractreminder'
    
    def get_success_url(self):
        return reverse_lazy('contracts:contract_detail', kwargs={'pk': self.object.contract.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, _('Lembrete criado com sucesso!'))
        return response
    
    def get_initial(self):
        initial = super().get_initial()
        contract_id = self.request.GET.get('contract')
        if contract_id:
            initial['contract'] = contract_id
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Novo Lembrete')
        return context


class ContractReminderUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = ContractReminder
    form_class = ContractReminderForm
    template_name = 'contracts/contractreminder_form.html'
    permission_required = 'contracts.change_contractreminder'
    
    def get_success_url(self):
        return reverse_lazy('contracts:contract_detail', kwargs={'pk': self.object.contract.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Lembrete atualizado com sucesso!'))
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Editar Lembrete')
        return context


class ContractReminderCompleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'contracts.change_contractreminder'
    
    def post(self, request, *args, **kwargs):
        reminder = get_object_or_404(ContractReminder, pk=kwargs['pk'])
        reminder.completed = not reminder.completed
        reminder.completed_at = timezone.now() if reminder.completed else None
        reminder.save()
        
        action = 'concluído' if reminder.completed else 'reaberto'
        messages.success(request, _(f'Lembrete {action} com sucesso!'))
        
        return redirect('contracts:contract_detail', pk=reminder.contract.pk)


class ContractReminderDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ContractReminder
    template_name = 'contracts/contractreminder_confirm_delete.html'
    permission_required = 'contracts.delete_contractreminder'
    
    def get_success_url(self):
        return reverse_lazy('contracts:contract_detail', kwargs={'pk': self.object.contract.pk})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Lembrete excluído com sucesso!'))
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Excluir Lembrete')
        return context


# Visualizações para Histórico de Contratos
class ContractHistoryListView(LoginRequiredMixin, ListView):
    model = ContractHistory
    template_name = 'contracts/contracthistory_list.html'
    context_object_name = 'history_entries'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = ContractHistory.objects.select_related(
            'contract', 'changed_by'
        ).order_by('-changed_at')
        
        # Filtro por contrato
        contract_id = self.request.GET.get('contract')
        if contract_id:
            queryset = queryset.filter(contract_id=contract_id)
        
        # Filtro por usuário
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(changed_by_id=user_id)
        
        # Filtro por data
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(changed_at__date__gte=date_from)
            
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(changed_at__date__lte=date_to)
        
        # Filtro por tipo de alteração
        change_type = self.request.GET.get('change_type')
        if change_type:
            queryset = queryset.filter(change_type=change_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Adicionar filtros atuais ao contexto
        context['contract_id'] = self.request.GET.get('contract', '')
        context['user_id'] = self.request.GET.get('user', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['change_type'] = self.request.GET.get('change_type', '')
        
        # Adicionar listas para os filtros
        context['contracts'] = Contract.objects.all().order_by('title')
        context['users'] = User.objects.filter(
            id__in=ContractHistory.objects.values('changed_by').distinct()
        ).order_by('first_name', 'last_name')
        
        # Opções de tipo de alteração
        context['change_types'] = ContractHistory.CHANGE_TYPES
        
        # Estatísticas
        
        return context
class ContractHistoryDetailView(LoginRequiredMixin, DetailView):
    model = ContractHistory
    template_name = 'contracts/contracthistory_detail.html'
    context_object_name = 'history_entry'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contract'] = self.object.contract

        changes = []
        for field, old_value, new_value in self.object.get_changes():
            # Traduzir o nome do campo, se aplicável
            field_display = dict(Contract._meta.get_field(field).flatchoices).get(
                field, field.replace('_', ' ').title()
            )

            # Formatar valores especiais
            if field.endswith('_at') or field.endswith('_date'):
                old_value = format_datetime(old_value) if old_value else None
                new_value = format_datetime(new_value) if new_value else None

            changes.append({
                'field': field_display,
                'old_value': old_value or _('(vazio)'),
                'new_value': new_value or _('(vazio)')
            })

        context['formatted_changes'] = changes
        context['title'] = _('Detalhes do Histórico')
        return context

# Visualizações para Anexos de Contrato
class ContractAttachmentListView(LoginRequiredMixin, ListView):
    model = ContractAttachment
    template_name = 'contracts/contractattachment_list.html'
    context_object_name = 'attachments'
    
    def get_queryset(self):
        self.contract = get_object_or_404(Contract, pk=self.kwargs['contract_pk'])
        return self.contract.attachments.all()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contract'] = self.contract
        context['title'] = _('Anexos do Contrato')
        return context


class ContractAttachmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = ContractAttachment
    form_class = ContractAttachmentForm
    template_name = 'contracts/contractattachment_form.html'
    permission_required = 'contracts.add_contractattachment'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['contract'] = get_object_or_404(Contract, pk=self.kwargs['contract_pk'])
        return kwargs
    
    def form_valid(self, form):
        form.instance.contract_id = self.kwargs['contract_pk']
        form.instance.uploaded_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, _('Anexo adicionado com sucesso!'))
        return response
    
    def get_success_url(self):
        return reverse('contracts:contract_attachment_list', kwargs={'contract_pk': self.kwargs['contract_pk']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contract'] = get_object_or_404(Contract, pk=self.kwargs['contract_pk'])
        context['title'] = _('Adicionar Anexo')
        return context


class ContractAttachmentDetailView(LoginRequiredMixin, DetailView):
    model = ContractAttachment
    template_name = 'contracts/contractattachment_detail.html'
    context_object_name = 'attachment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Detalhes do Anexo')
        return context


class ContractAttachmentDownloadView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        attachment = get_object_or_404(ContractAttachment, pk=kwargs['pk'])
        
        # Registrar o download no histórico
        ContractHistory.objects.create(
            contract=attachment.contract,
            changed_by=request.user,
            change_type='download',
            description=f'Anexo "{attachment.filename}" baixado',
            changes={}
        )
        
        # Configurar a resposta para download do arquivo
        response = FileResponse(attachment.file)
        response['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
        return response


class ContractAttachmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = ContractAttachment
    template_name = 'contracts/contractattachment_confirm_delete.html'
    permission_required = 'contracts.delete_contractattachment'
    
    def get_success_url(self):
        return reverse('contracts:contract_attachment_list', 
                     kwargs={'contract_pk': self.object.contract.pk})
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # Registrar a exclusão no histórico
        ContractHistory.objects.create(
            contract=self.object.contract,
            changed_by=request.user,
            change_type='delete',
            description=f'Anexo "{self.object.filename}" excluído',
            changes={}
        )
        
        # Excluir o arquivo físico
        self.object.file.delete(save=False)
        
        # Excluir o registro do banco de dados
        self.object.delete()
        
        messages.success(request, _('Anexo excluído com sucesso!'))
        return HttpResponseRedirect(success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Excluir Anexo')
        context['contract'] = self.object.contract
        return context


# Visualizações para Relatórios e Dashboards
class ContractDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'contracts/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Estatísticas gerais
        context['total_contracts'] = Contract.objects.count()
        context['active_contracts'] = Contract.objects.filter(
            end_date__gte=today,
            start_date__lte=today
        ).count()
        
        # Contratos a vencer nos próximos 30 dias
        context['expiring_soon'] = Contract.objects.filter(
            end_date__range=(today, today + timezone.timedelta(days=30))
        ).order_by('end_date')[:5]
        
        # Contratos vencidos
        context['expired_contracts'] = Contract.objects.filter(
            end_date__lt=today
        ).order_by('-end_date')[:5]
        
        # Contratos por tipo
        context['contracts_by_type'] = ContractType.objects.annotate(
            count=Count('contracts')
        ).filter(count__gt=0).order_by('-count')
        
        # Contratos por status
        context['contracts_by_status'] = ContractStatus.objects.annotate(
            count=Count('contracts')
        ).filter(count__gt=0).order_by('-count')
        
        # Lembretes pendentes
        context['pending_reminders'] = ContractReminder.objects.filter(
            completed=False,
            due_date__lte=today + timezone.timedelta(days=7)
        ).order_by('due_date')[:10]
        
        # Atividades recentes
        context['recent_activities'] = ContractHistory.objects.select_related(
            'contract', 'changed_by'
        ).order_by('-changed_at')[:10]
        
        context['title'] = _('Painel de Controle')
        return context


class ContractReportView(LoginRequiredMixin, TemplateView):
    template_name = 'contracts/reports/contract_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtros
        contract_type = self.request.GET.get('contract_type')
        status = self.request.GET.get('status')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Aplicar filtros
        queryset = Contract.objects.all()
        
        if contract_type:
            queryset = queryset.filter(contract_type_id=contract_type)
        
        if status:
            queryset = queryset.filter(status_id=status)
            
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
            
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        # Adicionar contagem de anexos e lembretes
        queryset = queryset.annotate(
            attachment_count=Count('attachments'),
            reminder_count=Count('reminders')
        )
        
        context['contracts'] = queryset.order_by('title')
        
        # Opções de filtro
        context['contract_types'] = ContractType.objects.all()
        context['statuses'] = ContractStatus.objects.all()
        
        # Manter os filtros ativos
        context['filters'] = {
            'contract_type': int(contract_type) if contract_type else '',
            'status': int(status) if status else '',
            'start_date': start_date or '',
            'end_date': end_date or ''
        }
        
        context['title'] = _('Relatório de Contratos')
        return context


class ContractExpirationReportView(LoginRequiredMixin, TemplateView):
    template_name = 'contracts/reports/expiration_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Contratos a vencer
        days = int(self.request.GET.get('days', 30))
        
        expiring_contracts = Contract.objects.filter(
            end_date__range=(today, today + timezone.timedelta(days=days))
        ).order_by('end_date')
        
        # Contratos vencidos
        expired_contracts = Contract.objects.filter(
            end_date__lt=today
        ).order_by('-end_date')
        
        # Agrupar por mês de vencimento
        expiration_dates = Contract.objects.filter(
            end_date__gte=today
        ).values_list('end_date', flat=True)
        
        expiration_by_month = {}
        for date in expiration_dates:
            month_year = date.strftime('%Y-%m')
            expiration_by_month[month_year] = expiration_by_month.get(month_year, 0) + 1
        
        context.update({
            'expiring_contracts': expiring_contracts,
            'expired_contracts': expired_contracts,
            'expiration_by_month': sorted(expiration_by_month.items()),
            'days': days,
            'title': _('Relatório de Vencimento de Contratos')
        })
        
        return context


class ContractValueReportView(LoginRequiredMixin, TemplateView):
    template_name = 'contracts/reports/value_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Valor total por tipo de contrato
        value_by_type = ContractType.objects.annotate(
            total_value=Sum('contracts__value'),
            contract_count=Count('contracts')
        ).filter(total_value__isnull=False)
        
        # Valor total por ano
        value_by_year = Contract.objects.exclude(
            start_date__isnull=True
        ).annotate(
            year=ExtractYear('start_date')
        ).values('year').annotate(
            total_value=Sum('value'),
            contract_count=Count('id')
        ).order_by('year')
        
        # Top contratos por valor
        top_contracts = Contract.objects.exclude(
            value__isnull=True
        ).order_by('-value')[:10]
        
        context.update({
            'value_by_type': value_by_type,
            'value_by_year': value_by_year,
            'top_contracts': top_contracts,
            'total_value': Contract.objects.aggregate(Sum('value'))['value__sum'] or 0,
            'title': _('Relatório de Valores de Contratos')
        })
        
        return context


def export_contracts_to_excel(request):
    """
    Exporta os contratos para um arquivo Excel.
    """
    # Criar o arquivo Excel
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    
    # Cabeçalhos
    headers = [
        'Número', 'Título', 'Tipo', 'Status', 'Valor',
        'Data Início', 'Data Término', 'Dias Restantes',
        'Partes', 'Anexos', 'Lembretes', 'Descrição'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Dados
    today = timezone.now().date()
    contracts = Contract.objects.select_related(
        'contract_type', 'status'
    ).prefetch_related('parties', 'attachments', 'reminders')
    
    for row_num, contract in enumerate(contracts, 2):
        days_left = (contract.end_date - today).days if contract.end_date else None
        
        ws.cell(row=row_num, column=1, value=contract.contract_number)
        ws.cell(row=row_num, column=2, value=str(contract))
        ws.cell(row=row_num, column=3, value=str(contract.contract_type) if contract.contract_type else '')
        ws.cell(row=row_num, column=4, value=str(contract.status) if contract.status else '')
        ws.cell(row=row_num, column=5, value=float(contract.value) if contract.value else 0)
        ws.cell(row=row_num, column=6, value=contract.start_date)
        ws.cell(row=row_num, column=7, value=contract.end_date)
        ws.cell(row=row_num, column=8, value=days_left)
        ws.cell(row=row_num, column=9, value=', '.join(str(p) for p in contract.parties.all()))
        ws.cell(row=row_num, column=10, value=contract.attachments.count())
        ws.cell(row=row_num, column=11, value=contract.reminders.count())
        ws.cell(row=row_num, column=12, value=contract.description)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar o arquivo
    wb.save(output)
    output.seek(0)
    
    # Criar a resposta
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=contratos_exportados.xlsx'
    
    return response


class ContractExportView(LoginRequiredMixin, View):
    """
    View para exportação de contratos em diferentes formatos (Excel, CSV, etc.)
    """
    def get(self, request, *args, **kwargs):
        # Obter parâmetros de filtro
        contract_type = request.GET.get('contract_type')
        status = request.GET.get('status')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        export_format = request.GET.get('format', 'excel')
        
        # Filtrar contratos
        queryset = Contract.objects.all()
        
        if contract_type:
            queryset = queryset.filter(contract_type_id=contract_type)
        if status:
            queryset = queryset.filter(status_id=status)
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        # Ordenar por data de término (mais próximos primeiro)
        queryset = queryset.order_by('end_date')
        
        # Exportar no formato solicitado
        if export_format == 'csv':
            return self.export_to_csv(queryset)
        else:  # Padrão para Excel
            return self.export_to_excel(queryset)
    
    def export_to_excel(self, queryset):
        """Exporta os contratos para um arquivo Excel"""
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Contratos"
        
        # Cabeçalhos
        headers = [
            'Número', 'Título', 'Tipo', 'Status', 'Valor',
            'Data Início', 'Data Término', 'Dias Restantes',
            'Descrição', 'Criado em', 'Atualizado em'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Dados
        today = timezone.now().date()
        
        for row_num, contract in enumerate(queryset, 2):
            days_left = (contract.end_date - today).days if contract.end_date else None
            
            ws.cell(row=row_num, column=1, value=contract.contract_number or '')
            ws.cell(row=row_num, column=2, value=str(contract))
            ws.cell(row=row_num, column=3, value=str(contract.contract_type) if contract.contract_type else '')
            ws.cell(row=row_num, column=4, value=str(contract.status) if contract.status else '')
            ws.cell(row=row_num, column=5, value=float(contract.value) if contract.value else 0)
            ws.cell(row=row_num, column=6, value=contract.start_date)
            ws.cell(row=row_num, column=7, value=contract.end_date)
            ws.cell(row=row_num, column=8, value=days_left)
            ws.cell(row=row_num, column=9, value=contract.description or '')
            ws.cell(row=row_num, column=10, value=contract.created_at)
            ws.cell(row=row_num, column=11, value=contract.updated_at)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar o arquivo
        wb.save(output)
        output.seek(0)
        
        # Criar a resposta
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=contratos_exportados.xlsx'
        return response
    
    def export_to_csv(self, queryset):
        """Exporta os contratos para um arquivo CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=contratos_exportados.csv'
        
        writer = csv.writer(response)
        writer.writerow([
            'Número', 'Título', 'Tipo', 'Status', 'Valor',
            'Data Início', 'Data Término', 'Dias Restantes',
            'Descrição', 'Criado em', 'Atualizado em'
        ])
        
        today = timezone.now().date()
        
        for contract in queryset:
            days_left = (contract.end_date - today).days if contract.end_date else None
            
            writer.writerow([
                contract.contract_number or '',
                str(contract),
                str(contract.contract_type) if contract.contract_type else '',
                str(contract.status) if contract.status else '',
                float(contract.value) if contract.value else 0,
                contract.start_date,
                contract.end_date,
                days_left,
                contract.description or '',
                contract.created_at,
                contract.updated_at
            ])
        
        return response


@require_http_methods(["GET"])
def check_contract_number(request):
    """
    Verifica se um número de contrato já existe no sistema.
    Retorna JSON indicando se o número já está em uso.
    """
    number = request.GET.get('number', '').strip()
    current_id = request.GET.get('current_id', '').strip()
    
    if not number:
        return JsonResponse({'available': False, 'message': 'Número de contrato não fornecido'}, status=400)
    
    # Verifica se já existe um contrato com este número
    from .models import Contract
    
    query = Contract.objects.filter(contract_number=number)
    
    # Se for uma atualização, exclui o contrato atual da verificação
    if current_id and current_id.isdigit():
        query = query.exclude(id=int(current_id))
    
    exists = query.exists()
    
    return JsonResponse({
        'available': not exists,
        'message': 'Número já está em uso' if exists else 'Número disponível'
    })


class ContractHistoryView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Visualização para exibir o histórico de alterações de um contrato específico.
    """
    model = ContractHistory
    template_name = 'contracts/contract_history.html'
    context_object_name = 'history_entries'
    paginate_by = 20
    permission_required = 'contracts.view_contracthistory'

    def get_queryset(self):
        """
        Retorna o histórico de alterações para o contrato especificado.
        """
        contract_id = self.kwargs.get('pk')
        return ContractHistory.objects.filter(contract_id=contract_id).order_by('-changed_at')

    def get_context_data(self, **kwargs):
        """
        Adiciona o contrato ao contexto para exibi��o no template.
        """
        context = super().get_context_data(**kwargs)
        context['contract'] = get_object_or_404(Contract, pk=self.kwargs.get('pk'))
        return context
